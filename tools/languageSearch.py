"""Python 'languageSearch.py' - Auto translation from exist file and fill into new file.

Based on Python 3

-*- coding: utf-8 -*-

Written by Jan (luojiayuan@hikvision.com).
"""

import sys
import os 
import json
#import imp
import xlrd
#import xlwt # xlwt 模块无法写入 xlsx表格，只能写xls表格，因此使用openpyxl模块操作
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill

def get_language_col(language, langline):
    # 传入语言行，查询语言在第几列
    if language in langline:
        return langline.index(language)
    else:
        return -1    #如果第一列无 语言，则默认选取第一列

def listdir(file_dir):   
    # 获取当前文件夹下的xls 或 xlsx文件
    file_path=[]   
    for dirpath,dirnames,files in os.walk(file_dir):  
        if dirpath == file_dir:
            for file in files: 
                if os.path.splitext(file)[1] == '.xls' or os.path.splitext(file)[1] == '.xlsx':  
                    file_path.append(file)  
    return file_path  

#imp.reload(sys)# Python2.5 初始化后删除了 sys.setdefaultencoding 方法，我们需要重新载入 
#filename = filename.decode('utf-8') #python 3 已经是utf-8

"""Info"""
print('\n    功能：将需要的翻译从多份文件中查找并整合到一份文件中.')
print('    资源：待翻译的表格，供检索的表格')
print('    输入：待翻译的表格名称，待翻译的语言，待检出的语言')
print('\n*****************************注意事项*****************************')
print('    请将需要翻译的表格以及需要检索的表格放到此执行程序路径下.')
print('    表格第一行为语言行，表示当前列的语言类型，如英语，法语等.')
print('    翻译的表格和检索的表格中，语言行的语言请保持一致，') 
print('    不要出现如 第一个表格写了 英语，第二个表格写了 English 情况.')
print('    检出表格中如有红色单元格，表示搜索不到此翻译.')
print('    检出表格如有多列，表示不同文件同一单词的翻译有冲突.')
print('******************************************************************\n')
# print('Please put [The file you need to translate] and [The file you need to search translation] to the Executable Program path.')
# print('And keep the Excel sheet fisrt line [language name line] to the same')
# print('   eg: If need translate English to France, ') 
# print('       [The file you need to translate] language line should contain English, ')
# print('       and [The file you need to search translation] language line should contain English and France. ')
# print('If you see red cell in output Excel, it means this word not in all searched files.')
# print('If you see multiple columns in output Excel, it means this word has conflict translation between diffrent files.')

"""初始化各种单元格格式"""
highlight = NamedStyle(name="highlight")
# 设置背景填充，红色
highlight.fill = PatternFill("solid", fgColor="ff0000")
# 设置边框，实线，黑色
highlight.border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'), bottom=Side(border_style="thin",color='000000'))

normalstyle = NamedStyle(name="normalstyle")
# 设置背景填充，白色
normalstyle.fill = PatternFill("solid", fgColor="ffffff") 
# 设置边框，实线，黑色
normalstyle.border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'), bottom=Side(border_style="thin",color='000000'))

indexstyle = NamedStyle(name="indexstyle")
# 设置背景填充，浅蓝色
indexstyle.fill = PatternFill("solid", fgColor="9bc2e6") 
# 设置边框，实线，黑色
indexstyle.border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'), bottom=Side(border_style="thin",color='000000'))


"""获取配置文件"""
filename = 'config.json'

strSrcfile = 'Source File'
strSrcLang = 'Source Language'
strDstLang = 'Dest Language'
new_dict = {}

try:
    with open(filename, 'r') as f_obj:
        new_dict = json.load(f_obj)
except FileNotFoundError:
    print('Creat config.json')

"""获取待翻译的文件"""
if strSrcfile in new_dict:
    need_trans = new_dict[strSrcfile]
    print('需要翻译的文件 [%s] ' %need_trans)
else:
    need_trans = input('请输入要翻译的文件(eg: translate.xlsx):\n') # Please input the file you need to translate 

"""打开读取的Excel, 读取需要翻译的英文"""
try:
    rd_need_book = xlrd.open_workbook(need_trans) # 打开xls文件
    rd_need_sheet = rd_need_book.sheets()[0] # 打开第一个sheet
except FileNotFoundError:
    print('File [%s] not exist' %need_trans)
    input('Please input any key to exit!')
    sys.exit(0)

"""获取被翻译的语言"""
if strSrcLang in new_dict:
    source_lang = new_dict[strSrcLang]
    print('需要翻译 [%s] ' %source_lang)
else:
    source_lang = input('请输入要翻译的语言(eg: English):\n') # Please input the language you need to translate 

"""获取需要翻译成的语言"""
if strDstLang in new_dict:
    need_lang = new_dict[strDstLang]
    print('需要翻译 [%s] \n' %need_lang)
else:    
    need_lang = input('请输入要翻译成的语言(eg: France):\n') # Please input the language you need to translate to 

save_file_name = need_lang + '_translation.xlsx'

"""获取待检索的文件"""
source_list = listdir('./') 
source_list.remove(need_trans)
print('The file you provide for searching:')
print(source_list)

if rd_need_sheet.nrows == 0:
    print('File [%s] is empty' %need_trans)
    input('Please input any key to exit!')
    sys.exit(0)   

trans_en_line = rd_need_sheet.row_values(0, 0, None) # 获得待翻译表格第一行语言行所有数据
trans_en_col_index = get_language_col(source_lang, trans_en_line) # 获取待翻译表格中 待翻译语言 的列索引
if trans_en_col_index == -1:
    print('File [%s] is not including [%s] you need !' %(need_trans, source_lang))
    input('Please input any key to exit!')
    sys.exit(0)   

"""打开写入的Excel"""
wt_book = Workbook()
wt_sheet = wt_book.active
FIRST_COL = 1 #索引从1开始计算

for sourcefile in source_list:
    """打开读取的Excel，用于检索"""
    try:
        rd_source_book = xlrd.open_workbook(sourcefile) # 打开xls文件
        rd_source_sheet = rd_source_book.sheets()[0] # 打开第一个sheet
    except FileNotFoundError:
        print('File [%s] is not exist' %sourcefile)
        continue

    if rd_source_sheet.nrows == 0:
        print('File [%s] is empty' %sourcefile)
        continue

    search_en_line = rd_source_sheet.row_values(0, 0, None) # 获得第一行所有数据
    search_en_index = get_language_col(source_lang, search_en_line) # 获取存在待搜索表格中 待翻译语言 的列索引
    language_col_index = get_language_col(need_lang, search_en_line) # 获取存在待搜索表格中 所需语言 的列索引

    cols_find = rd_source_sheet.col_values(search_en_index, 0, None) # 获得存在待搜索表格中 待翻译语言 所在列的所有数据

    if search_en_index == -1:
        print('File [%s] is not including [%s] you need !' %(sourcefile, source_lang))
        continue
    if language_col_index == -1:
        print('File [%s] is not including [%s] you need !' %(sourcefile, need_lang))
        continue

    i = 0
    for i in range(rd_need_sheet.nrows):
        if rd_need_sheet.cell(i, trans_en_col_index).value in set(cols_find): # 待翻译语言 是否存在于搜索列中
            if rd_need_sheet.cell(i, trans_en_col_index).value.startswith('[') and rd_need_sheet.cell(i, trans_en_col_index).value.endswith(']'):
                # 如果是索引行，拷贝
                wt_sheet.cell(i+1, FIRST_COL).style = indexstyle # 设置索引行单元格格式
            else:
                wt_sheet.cell(i+1, FIRST_COL).style = normalstyle # 设置普通行单元格格式
            nPos=cols_find.index( rd_need_sheet.cell(i, trans_en_col_index).value ) # 查找 语言 在第几行 
            if wt_sheet.cell(i+1, FIRST_COL).value != None and wt_sheet.cell(i+1, FIRST_COL).value != rd_source_sheet.cell(nPos, language_col_index).value:         
                # 翻译在多个表格中有 分歧，追加到另一列，最多追加2列 
                if wt_sheet.cell(i+1, FIRST_COL+1).value == None:
                    wt_sheet.cell(i+1, FIRST_COL+1, value=rd_source_sheet.cell(nPos, language_col_index).value)
                elif wt_sheet.cell(i+1, FIRST_COL+1).value != rd_source_sheet.cell(nPos, language_col_index).value:
                    wt_sheet.cell(i+1, FIRST_COL+2, value=rd_source_sheet.cell(nPos, language_col_index).value)
            else:
                wt_sheet.cell(i+1, FIRST_COL, value=rd_source_sheet.cell(nPos, language_col_index).value) # 设置对应的语言
        else:
            # 未找到行填充背景色
            if rd_need_sheet.cell(i, trans_en_col_index).value.startswith('[') and rd_need_sheet.cell(i, trans_en_col_index).value.endswith(']'):
                # 如果是索引行，拷贝
                wt_sheet.cell(i+1, FIRST_COL).style = indexstyle
                wt_sheet.cell(i+1, FIRST_COL, value=rd_need_sheet.cell(i, trans_en_col_index).value) # 设置对应的语言
            else:
                if wt_sheet.cell(i+1, FIRST_COL).value == None:
                    wt_sheet.cell(i+1, FIRST_COL).style = highlight # 不存在的设置单元格背景为红色

        i += 1  

    print('Translate successfully from file [%s] and has already saved to [%s].' %(sourcefile, save_file_name))

"""保存写入的Excel"""
wt_book.save(save_file_name) 

config_data = {}
config_data['Source File'] = need_trans    
config_data['Source Language'] = source_lang
config_data['Dest Language'] = need_lang
with open(filename, 'w') as f_obj:
    json.dump(config_data, f_obj)
    
input('Please input any key to exit!')